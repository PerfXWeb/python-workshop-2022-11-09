#First things first we need our packages, which will enable us to communicate with differente files
import os
import openpyxl 
import matplotlib.pyplot as plt
import numpy as np

#In order to communicate with Excel files we do need their path.... It is assumed that the excel files will be in the same repository as the Python file
folder_name = os.getcwd()
filename = os.path.join(folder_name, "stocks_unclean.xlsx")#this will just connect the current working directory with the second argument

if os.path.exists(filename) == False:
    print(f"{filename} file does not exist!")
    exit(1)
# To open the workbook we will use our package openpyxl
# This will return us a workbook 
wb = openpyxl.load_workbook(filename)

#Get the different sheets of the Excel file
sheet1 = wb['INTC']
sheet2= wb['TSM']

#This function will be called later to validate the data read
#Here the price examined wether it is a number and in some kind of realistic frame
#and the format of the date is checked
def validation(date, price):
    try:
        price = float(price)
    except:
        print(f"Error on row {i}: Price is not Numeric ({price})")
        return False
    if price<=0 and price>=10E100:
        print(f"Error on row {i}: Price is nuts ({price})")
        return False
    if 'datetime' not in str(type(date)) or date=="":
        print(f"Error on row {i}: Date is not valid {date}")
        return False
    return True


#Now we will extract the content of the single cells and place it into a dictionary
dict1 = {}
for i in reversed(range(2, sheet1.max_row)):
    #remember we splittet the workbook into to different sheets. We are extracting cell content of the first sheet 
    date = sheet1.cell(row = i, column = 1).value
    price = sheet1.cell(row = i, column = 5).value
    
    #After we extracted the content we have to make sure it's valid 
    if validation(date, price) is False:
        sheet1.delete_rows(i)
    else:
        dict1[date] = price 
#print(dict1)

#The same steps as above are used to extract and check the content of the cells
dict2 = {}
for i in reversed(range(2, sheet2.max_row)):
    date = sheet2.cell(row = i, column = 1).value
    price = sheet2.cell(row = i, column = 5).value
    if validation(date, price) is False:
        sheet2.delete_rows(i)
    else:
        dict2[date] = price
#print(dict2)

#It might occur that the stock prices are not recorded with the same time stamps all along
#Therefore we will compare the entries with their timestamps
values = []
for key in dict1:
    if key in dict2:
        #If the same timestamps are found in both directories they will be stored in a 2 dimensional array 
        values.append([key, dict1[key], dict2[key]])
    else:
        print(f"Skipped date {key}")
#print(dict)

values = np.array(values)

#Now the data is 
# 1. validated 
# 2. made compatible
# This means the data is ready for comparison 

# We could write them in a new results sheet
wb.create_sheet('Results')
for e,i in enumerate(values):
    wb['Results'].cell(row = e+1, column = 1).value = i[0]
    wb['Results'].cell(row = e+1, column = 2).value = i[1]
    wb['Results'].cell(row = e+1, column = 3).value = i[2]
wb.save('stocks_cleaned.xlsx')

#We could plot them via using matplotlib
plt.plot(values[:,0], values[:,1])
plt.plot(values[:,0], values[:,2])
plt.show()

#We could also do some far more exiting stuff like a scattering plot
#... with just a few more lines of code
plt.scatter(values[:,1], values[:,2])
plt.title('Scatter plot')
plt.xlabel('INTL')
plt.ylabel('TSMC')
plt.show()

